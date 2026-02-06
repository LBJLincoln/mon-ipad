#!/usr/bin/env python3
"""
Generate comprehensive RAG Benchmark Datasets reference document (XLSX)
30 datasets used by Anthropic, Google, Meta, OpenAI, Microsoft et al.
for evaluating RAG systems — sourced from academic papers 2023-2026.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ============================================================
# STYLES
# ============================================================
header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
cat_font = Font(name="Calibri", bold=True, size=11, color="1F4E79")
cat_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
wrap = Alignment(wrap_text=True, vertical="top")
thin_border = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)

# ============================================================
# SHEET 1: 30 RAG BENCHMARK DATASETS
# ============================================================
ws1 = wb.active
ws1.title = "30 RAG Datasets"

headers = [
    "N°", "Dataset", "Catégorie RAG", "Type de Test", "Taille",
    "Source/Papier", "Utilisé par", "URL HuggingFace / GitHub",
    "Format Q&A", "Exemple Question", "Exemple Réponse Attendue",
    "Métriques Clés", "Pertinence pour ton RAG n8n"
]

# Column widths
col_widths = [4, 18, 16, 18, 12, 28, 22, 36, 18, 32, 32, 22, 28]

for i, (h, w) in enumerate(zip(headers, col_widths), 1):
    cell = ws1.cell(row=1, column=i, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = wrap
    cell.border = thin_border
    ws1.column_dimensions[get_column_letter(i)].width = w

# 30 datasets
datasets = [
    # === SINGLE-HOP QA (Retrieval basique) ===
    ["CAT", "🔹 SINGLE-HOP QA — Retrieval basique"],
    [1, "Natural Questions (NQ)", "Single-Hop QA", "Retrieval + Generation", "~307k Q&A",
     "Kwiatkowski et al. 2019 (Google)", "Google, Anthropic, Meta, OpenAI",
     "google-research-datasets/nq_open",
     "question → short/long answer", "When was the last time the Chicago Cubs won the World Series?",
     "2016", "EM, F1, Recall@k", "Test basique: ton RAG retrouve-t-il la bonne info dans un seul document?"],
    [2, "TriviaQA", "Single-Hop QA", "Retrieval + Reasoning", "~95k Q&A",
     "Joshi et al. 2017 (UW)", "Google, Anthropic, Meta, HippoRAG",
     "mandarjoshi/trivia_qa",
     "question → answer + evidence", "Which American-born Dadaist is associated with both New York and Paris?",
     "Man Ray", "EM, F1", "Test de recall: questions factuelles nécessitant un seul passage pertinent"],
    [3, "SQuAD 2.0", "Single-Hop QA", "Reading Comprehension", "~150k Q&A",
     "Rajpurkar et al. 2018 (Stanford)", "Google, OpenAI, Meta",
     "rajpurkar/squad_v2",
     "context + question → answer span", "What is the capital of Australia?",
     "Canberra", "EM, F1", "Test de compréhension extractive: le LLM extrait-il la bonne span?"],
    [4, "PopQA", "Single-Hop QA", "Entity-Centric QA", "~14k Q&A",
     "Mallen et al. 2023", "HippoRAG, SELF-RAG, Anthropic",
     "akariasai/PopQA",
     "question → entity answer", "Who is the director of the movie Inception?",
     "Christopher Nolan", "Accuracy", "Test entités: les entités rares nécessitent-elles le RAG vs mémoire LLM?"],
    [5, "WebQuestions", "Single-Hop QA", "Web-sourced QA", "~6.6k Q&A",
     "Berant et al. 2013 (Stanford)", "Google, Meta",
     "Stanford WebQuestions",
     "question → answer", "What language do they speak in Thailand?",
     "Thai", "Accuracy, F1", "Test questions simples issues de recherches web réelles"],

    # === MULTI-HOP QA (Raisonnement multi-documents) ===
    ["CAT", "🔹 MULTI-HOP QA — Raisonnement multi-documents"],
    [6, "HotpotQA", "Multi-Hop QA", "Multi-doc Reasoning", "~113k Q&A",
     "Yang et al. 2018 (CMU+Stanford)", "Google, Anthropic, Meta, IBM, RAGBench",
     "hotpotqa/hotpot_qa",
     "question → answer + supporting facts", "Were Scott Derrickson and Ed Wood of the same nationality?",
     "Yes (both American)", "EM, F1, Sup. F1", "CRITIQUE: teste si ton RAG combine info de 2+ documents"],
    [7, "MuSiQue", "Multi-Hop QA", "Compositional Reasoning", "~25k Q&A",
     "Trivedi et al. 2022 (UNC)", "HippoRAG, Anthropic, Google",
     "StonyBrookNLP/musique",
     "multi-hop question → decomposed answer", "What is the population of the country where the Eiffel Tower is located?",
     "~67 million (France)", "EM, F1", "Test de raisonnement compositionnel: chaîne de 2-4 hops"],
    [8, "2WikiMultiHopQA", "Multi-Hop QA", "Cross-doc Reasoning", "~192k Q&A",
     "Ho et al. 2020", "HippoRAG, Meta",
     "IIRC/2wikimultihopqa",
     "question → answer + evidence chain", "Who is older, the director of Jaws or the director of ET?",
     "Steven Spielberg (même personne)", "EM, F1", "Test cross-document: raisonnement entre articles Wikipedia"],
    [9, "MultiHop-RAG", "Multi-Hop QA", "News-based Multi-hop", "~2.5k Q&A",
     "Tang et al. 2024", "Meta, divers",
     "yixuantt/MultiHopRAG",
     "multi-hop question → answer from news", "What company acquired the AI startup that launched product X?",
     "Varies", "Accuracy, F1", "Test multi-hop sur articles d'actualité (pas Wikipedia)"],

    # === RAG BENCHMARKS COMPLETS ===
    ["CAT", "🔹 RAG BENCHMARKS COMPLETS — Évaluation end-to-end"],
    [10, "CRAG (Meta KDD Cup)", "RAG Benchmark", "Comprehensive RAG", "~4.4k Q&A",
     "Yang et al. 2024 (Meta)", "Meta, KDD Cup participants",
     "github.com/facebookresearch/CRAG",
     "question → answer (web+KG)", "What was the closing price of AAPL on March 15, 2024?",
     "Dynamic (from web)", "Perfect/Accept/Miss/Incorrect", "RÉFÉRENCE: benchmark RAG complet avec web + knowledge graphs"],
    [11, "FRAMES (Google)", "RAG Benchmark", "Multi-hop RAG", "824 Q&A",
     "Krishna et al. 2024 (Google)", "Google, DeepMind",
     "google/frames-benchmark",
     "multi-hop Q → answer + wiki articles", "How many years separate the founding of MIT and Stanford?",
     "6 years (1861 vs 1885)", "Accuracy, Retrieval Precision", "CRITIQUE: test officiel Google pour RAG multi-hop avec raisonnement"],
    [12, "RAGBench", "RAG Benchmark", "Industry RAG (12 sous-datasets)", "~100k examples",
     "Friel et al. 2024 (Galileo)", "Galileo, divers",
     "rungalileo/ragbench",
     "question + context → answer + TRACe labels", "What is the treatment for condition X?",
     "Varies per domain", "TRACe: Relevance, Utilization, Completeness, Hallucination", "ESSENTIEL: 12 domaines industrie (legal, medical, finance, tech)"],
    [13, "RGB", "RAG Benchmark", "4 RAG Abilities", "~400 Q&A (EN+CN)",
     "Chen et al. 2023", "divers chercheurs",
     "github (chen-jiawei/RGB)",
     "question + noisy context → answer", "Based on the context, what is X? [avec documents bruités]",
     "Varies", "Noise Robustness, Negative Rejection, Info Integration, Counterfactual", "Test des 4 capacités fondamentales du générateur RAG"],
    [14, "CRUD-RAG", "RAG Benchmark", "CRUD Operations", "~4 task types",
     "Lyu et al. 2024", "divers",
     "github (IAAR-Shanghai/CRUD-RAG)",
     "CRUD tasks on knowledge base", "Create/Read/Update/Delete operations sur KB",
     "Varies per task", "Task-specific accuracy", "Test des opérations CRUD sur ta base de connaissances"],

    # === RETRIEVAL-ONLY (BEIR & IR) ===
    ["CAT", "🔹 RETRIEVAL-ONLY — Évaluation du retriever"],
    [15, "MS MARCO", "Retrieval", "Passage Retrieval", "~8.8M passages",
     "Nguyen et al. 2016 (Microsoft)", "Microsoft, Google, Anthropic, tous",
     "microsoft/ms_marco",
     "query → relevant passages", "What is the boiling point of water?",
     "Passage contenant '100°C at sea level'", "MRR@10, NDCG@10, Recall@1000", "Test de retrieval à grande échelle sur requêtes Bing réelles"],
    [16, "BEIR (18 datasets)", "Retrieval", "Zero-shot IR", "18 datasets variés",
     "Thakur et al. 2021", "Google, Meta, tous les retriever papers",
     "BeIR benchmark",
     "query → relevant documents", "Varies across 18 datasets",
     "Varies", "NDCG@10", "Benchmark de référence pour évaluer ton retriever (dense/sparse/hybrid)"],
    [17, "NoMIRACL", "Retrieval", "Robustness to Errors", "Multi-lingual",
     "Thakur et al. 2023", "Waterloo, Google",
     "Project NoMIRACL",
     "query + irrelevant docs → abstain/answer", "Question with only non-relevant retrieved docs",
     "I don't know / abstain", "Hallucination Rate, Error Rate", "Test CRITIQUE: ton RAG dit-il 'je ne sais pas' quand aucun doc pertinent?"],

    # === DOMAIN-SPECIFIC ===
    ["CAT", "🔹 DOMAIN-SPECIFIC — Évaluation par domaine"],
    [18, "PubMedQA", "Domain: Medical", "Biomedical QA", "~273k Q&A",
     "Jin et al. 2019", "MedRAG, Anthropic, Google",
     "qiaojin/PubMedQA",
     "question + abstract → yes/no/maybe", "Does vitamin D supplementation reduce cancer risk?",
     "Maybe (based on abstract)", "Accuracy", "Test RAG médical: raisonnement sur abstracts scientifiques"],
    [19, "FinQA", "Domain: Finance", "Financial Reasoning", "~8.2k Q&A",
     "Chen et al. 2021", "RAGBench, divers",
     "dreamerdeo/finqa",
     "question + financial table → numerical answer", "What is the gross margin for 2022?",
     "42.3% (calculated from table)", "Accuracy, Program Accuracy", "Test RAG finance: extraction + calcul depuis tableaux financiers"],
    [20, "CUAD", "Domain: Legal", "Contract Understanding", "~13k Q&A",
     "Hendrycks et al. 2021", "RAGBench, LegalBench",
     "CUADv1",
     "question about contract clause → answer", "Does this contract contain an exclusivity clause?",
     "Yes, Section 4.2 states...", "EM, AUPR", "Test RAG juridique: compréhension de contrats longs"],
    [21, "LegalBench-RAG", "Domain: Legal", "Legal RAG", "Varies",
     "ZeroEntropy 2024", "ZeroEntropy, divers",
     "zeroentropy-ai/legalbenchrag",
     "legal question + corpus → answer", "What are the limitations of liability in this agreement?",
     "Varies", "Accuracy, Retrieval Precision", "Benchmark RAG spécialisé juridique"],
    [22, "CovidQA", "Domain: Medical", "COVID-19 QA", "~2k Q&A",
     "Möller et al. 2020", "RAGBench",
     "RAGBench subset",
     "question about COVID → answer from papers", "What is the incubation period of COVID-19?",
     "2-14 days (typically 5 days)", "F1, EM", "Test RAG sur corpus scientifique COVID"],
    [23, "TechQA", "Domain: Technical", "IT Support QA", "~800 Q&A",
     "Castelli et al. 2020 (IBM)", "RAGBench, IBM",
     "RAGBench subset",
     "technical question → answer from docs", "How to resolve error code 0x80070005?",
     "Run as administrator...", "F1, EM", "Test RAG technique: résolution de problèmes IT"],

    # === LONG-FORM & COMPLEX ===
    ["CAT", "🔹 LONG-FORM & COMPLEX — Génération longue et complexe"],
    [24, "ASQA", "Long-form QA", "Ambiguous QA", "~6.3k Q&A",
     "Stelmakh et al. 2022", "RAG-bench, divers",
     "din0s/asqa",
     "ambiguous question → long-form answer", "Who played Batman? [multiple actors across years]",
     "Adam West (1966), Michael Keaton (1989), ...", "ROUGE-L, D-F1, Citation Precision", "Test réponses longues: gestion de l'ambiguïté et citations multiples"],
    [25, "ELI5", "Long-form QA", "Explanatory QA", "~272k Q&A",
     "Fan et al. 2019 (Facebook)", "RAG-bench, Meta",
     "eli5 (Reddit)",
     "question → long explanation", "ELI5: How does WiFi work?",
     "Long explanatory answer", "ROUGE-L", "Test génération longue: explications accessibles"],
    [26, "NarrativeQA", "Long-form QA", "Story Comprehension", "~46.8k Q&A",
     "Kočiský et al. 2018 (DeepMind)", "Google, HippoRAG",
     "deepmind/narrativeqa",
     "story + question → free-form answer", "What motivates the main character to leave home?",
     "Free-form narrative answer", "ROUGE-L, BLEU, METEOR", "Test compréhension longue: documents narratifs complets"],

    # === EVALUATION FRAMEWORKS ===
    ["CAT", "🔹 EVALUATION FRAMEWORKS — Frameworks d'évaluation RAG"],
    [27, "RAGAS Synthetic", "Eval Framework", "Synthetic Test Generation", "Generatable",
     "Es et al. 2023", "Utilisé par tous (framework)",
     "explodinggradients/ragas",
     "auto-generated Q&A from your docs", "Auto-generated from your corpus",
     "Auto-generated", "Faithfulness, Answer Relevancy, Context Precision/Recall", "ESSENTIEL: génère des Q&A de test directement depuis TES documents"],
    [28, "RAGChecker", "Eval Framework", "Fine-grained RAG Eval", "N/A (tool)",
     "NeurIPS 2024", "Amazon, divers",
     "amazon-science/RAGChecker",
     "claim-level evaluation", "N/A (evaluates your RAG output)",
     "N/A", "Claim Recall, Precision, Hallucination", "Framework d'évaluation claim-level (plus précis que RAGAS)"],
    [29, "MTRAG (IBM)", "RAG Benchmark", "Multi-Turn RAG", "~1.2k conversations",
     "IBM Research 2025", "IBM",
     "ibm/mt-rag-benchmark",
     "multi-turn conversation → answers", "Follow-up: And what about the second quarter?",
     "Context-dependent", "Turn Accuracy, Conversation Coherence", "Test RAG conversationnel multi-tours (crucial pour chatbots)"],
    [30, "OmniEval", "Eval Framework", "Multi-dimensional RAG Eval", "Comprehensive",
     "RUC-NLPIR 2024", "Renmin University",
     "RUC-NLPIR/OmniEval",
     "multi-dimensional evaluation", "Evaluates across dimensions",
     "N/A", "Multiple dimensions", "Évaluation multi-dimensionnelle couvrant tous les aspects RAG"],
]

row = 2
for d in datasets:
    if d[0] == "CAT":
        # Category separator row
        ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(headers))
        cell = ws1.cell(row=row, column=1, value=d[1])
        cell.font = cat_font
        cell.fill = cat_fill
        cell.alignment = Alignment(vertical="center")
        cell.border = thin_border
        for c in range(2, len(headers)+1):
            ws1.cell(row=row, column=c).fill = cat_fill
            ws1.cell(row=row, column=c).border = thin_border
        row += 1
    else:
        for i, val in enumerate(d, 1):
            cell = ws1.cell(row=row, column=i, value=val)
            cell.alignment = wrap
            cell.border = thin_border
            cell.font = Font(name="Calibri", size=10)
        row += 1

ws1.freeze_panes = "A2"
ws1.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{row-1}"

# ============================================================
# SHEET 2: MAPPING DATASETS → COMPOSANTS RAG
# ============================================================
ws2 = wb.create_sheet("Mapping RAG Components")

mapping_headers = ["Composant RAG testé", "Datasets recommandés", "Quoi mesurer", "Nœud n8n concerné"]
for i, h in enumerate(mapping_headers, 1):
    cell = ws2.cell(row=1, column=i, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = wrap
    cell.border = thin_border

ws2.column_dimensions["A"].width = 24
ws2.column_dimensions["B"].width = 36
ws2.column_dimensions["C"].width = 36
ws2.column_dimensions["D"].width = 28

mapping_data = [
    ["Chunking (Semantic)", "SQuAD 2.0, NarrativeQA, CUAD", "Chunk contient-il l'info? Chunk boundaries OK?", "Semantic Chunker V4"],
    ["Embedding Model", "MS MARCO, BEIR, NQ", "Recall@k, NDCG@10 sur retrieval", "Generate Embeddings V4"],
    ["BM25 Sparse Search", "MS MARCO, BEIR, TriviaQA", "Recall@k BM25 vs Dense vs Hybrid", "BM25 Sparse Vector Generator"],
    ["Hybrid Search (Dense+Sparse)", "MS MARCO, HotpotQA, NQ", "Recall gain hybrid vs dense-only", "Pinecone Hybrid Upsert"],
    ["Reranker", "MS MARCO, BEIR, HotpotQA", "MRR@10 after reranking", "Reranker (Qwen3-Reranker)"],
    ["Contextual Retrieval", "HotpotQA, MuSiQue, PopQA", "Retrieval accuracy avec/sans contexte", "Contextual Retrieval LLM"],
    ["Q&A Hypothétiques", "NQ, TriviaQA, PopQA", "Recall boost avec HyDE", "Q&A Generator V4"],
    ["Générateur (LLM)", "RGB, ASQA, ELI5, RAGAS", "Faithfulness, Hallucination Rate", "Agent / LLM Response"],
    ["Multi-Hop Reasoning", "HotpotQA, MuSiQue, 2Wiki, FRAMES", "Multi-hop accuracy, Supporting facts F1", "Orchestrator Agent"],
    ["Noise Robustness", "RGB, NoMIRACL", "Performance avec docs non pertinents", "Orchestrator + Retrieval"],
    ["Negative Rejection", "RGB, NoMIRACL, CRAG", "Le LLM dit-il 'je ne sais pas' quand il faut?", "Agent System Prompt"],
    ["Domain: Medical", "PubMedQA, CovidQA, MedRAG", "Accuracy médicale", "RAG médical"],
    ["Domain: Legal", "CUAD, LegalBench-RAG", "Extraction clauses, précision juridique", "RAG juridique"],
    ["Domain: Finance", "FinQA, TAT-QA", "Calculs financiers, extraction tableaux", "RAG finance"],
    ["Domain: Technical", "TechQA, EManual", "Résolution problèmes techniques", "RAG technique"],
    ["Multi-Turn Conversation", "MTRAG, CRAG Task 3", "Cohérence conversationnelle", "Orchestrator n8n"],
    ["End-to-End RAG", "CRAG, FRAMES, RAGBench", "Score global Perfect/Acceptable/Incorrect", "Pipeline complet"],
]

for r, row_data in enumerate(mapping_data, 2):
    for c, val in enumerate(row_data, 1):
        cell = ws2.cell(row=r, column=c, value=val)
        cell.alignment = wrap
        cell.border = thin_border
        cell.font = Font(name="Calibri", size=10)

ws2.freeze_panes = "A2"

# ============================================================
# SHEET 3: PLAN DE TEST n8n
# ============================================================
ws3 = wb.create_sheet("Plan Test n8n")

test_headers = [
    "Phase", "Workflow n8n", "Datasets utilisés", "Batch Size",
    "Métriques", "Seuil Acceptable", "Monitoring"
]
for i, h in enumerate(test_headers, 1):
    cell = ws3.cell(row=1, column=i, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = wrap
    cell.border = thin_border

ws3.column_dimensions["A"].width = 18
ws3.column_dimensions["B"].width = 30
ws3.column_dimensions["C"].width = 28
ws3.column_dimensions["D"].width = 12
ws3.column_dimensions["E"].width = 28
ws3.column_dimensions["F"].width = 16
ws3.column_dimensions["G"].width = 24

test_plan = [
    ["Phase 1: Ingestion", "WF1: Dataset Ingestion Pipeline", "NQ (1k sample), SQuAD (1k), PopQA (1k)", "50 items/batch",
     "Chunks created, Embeddings count, Ingestion time/doc", "< 30s/doc, 0 errors", "OTEL traces, Supabase logs"],
    ["Phase 1: Ingestion", "WF1: Dataset Ingestion Pipeline", "HotpotQA (1k), MS MARCO (5k)", "50 items/batch",
     "Vector count, BM25 index size", "Consistent counts", "Pinecone dashboard, ES stats"],
    ["Phase 2: Retrieval", "WF2: Retrieval Benchmark", "NQ, TriviaQA, PopQA (single-hop)", "20 queries/batch",
     "Recall@5, Recall@10, MRR@10", "Recall@10 > 0.75", "Google Sheets log"],
    ["Phase 2: Retrieval", "WF2: Retrieval Benchmark", "HotpotQA, MuSiQue (multi-hop)", "20 queries/batch",
     "Recall@10, Supporting Facts F1", "Recall@10 > 0.60", "Google Sheets log"],
    ["Phase 2: Retrieval", "WF2: Retrieval Benchmark", "MS MARCO (passage retrieval)", "20 queries/batch",
     "NDCG@10, MRR@10", "NDCG@10 > 0.40", "Google Sheets log"],
    ["Phase 3: Génération", "WF3: Generation Benchmark", "RGB (4 abilities)", "10 queries/batch",
     "Noise Robustness, Negative Rejection, Info Integration, CF Robustness", "Score > 0.70 chaque", "n8n Evaluation dashboard"],
    ["Phase 3: Génération", "WF3: Generation Benchmark", "ASQA, ELI5 (long-form)", "5 queries/batch",
     "ROUGE-L, Faithfulness (RAGAS), Citation Precision", "Faithfulness > 0.80", "n8n Evaluation dashboard"],
    ["Phase 4: End-to-End", "WF4: E2E RAG Benchmark", "FRAMES (824 Q&A complètes)", "10 queries/batch",
     "Accuracy, Retrieval Precision, Reasoning correctness", "Accuracy > 0.50", "n8n Eval + Google Sheets"],
    ["Phase 4: End-to-End", "WF4: E2E RAG Benchmark", "CRAG (sample 500)", "10 queries/batch",
     "Perfect/Acceptable/Missing/Incorrect rates", "Perfect+Accept > 60%", "n8n Eval + Google Sheets"],
    ["Phase 5: Domain", "WF5: Domain-Specific Benchmarks", "PubMedQA, CovidQA", "10 queries/batch",
     "Accuracy, Faithfulness", "Accuracy > 0.65", "n8n Eval"],
    ["Phase 5: Domain", "WF5: Domain-Specific Benchmarks", "FinQA, CUAD, LegalBench-RAG", "10 queries/batch",
     "Accuracy, Extraction precision", "Accuracy > 0.55", "n8n Eval"],
    ["Phase 6: Orchestrator", "WF6: Orchestrator Stress Test", "MTRAG (multi-turn), CRAG Task 3", "5 conversations/batch",
     "Turn accuracy, Tool selection accuracy, Coherence", "Turn accuracy > 0.70", "n8n Eval + OTEL"],
    ["Phase 7: Robustness", "WF7: Robustness Testing", "NoMIRACL, RGB (counterfactual)", "20 queries/batch",
     "Hallucination rate, Abstention rate", "Hallucination < 15%", "n8n Eval dashboard"],
    ["Phase 8: Regression", "WF8: Regression Suite (weekly)", "Mix: 50 NQ + 50 HotpotQA + 50 FRAMES + 50 RGB", "20/batch",
     "All key metrics aggregated", "No regression > 5% from baseline", "Automated n8n + Slack alerts"],
]

for r, row_data in enumerate(test_plan, 2):
    for c, val in enumerate(row_data, 1):
        cell = ws3.cell(row=r, column=c, value=val)
        cell.alignment = wrap
        cell.border = thin_border
        cell.font = Font(name="Calibri", size=10)

ws3.freeze_panes = "A2"

# ============================================================
# SAVE
# ============================================================
output_path = "/home/claude/RAG_30_Benchmark_Datasets_Reference.xlsx"
wb.save(output_path)
print(f"✅ Created: {output_path}")
